# -*- coding: utf-8 -*-

import base64
import pandas as pd
import io
import os

from pytz import timezone
from datetime import datetime
from openpyxl import load_workbook

from odoo import api, fields, models, tools, modules, _
from odoo.exceptions import UserError, ValidationError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'
    _order = 'create_date desc'
    
    estructure_id = fields.Many2one('hr.payroll.structure', 'Estructure', required=True)
    contracting_regime = fields.Selection([
        ('02', 'Wages and salaries'),
        ('05', 'Free'),
        ('08', 'Assimilated commission agents'),
        ('09', 'Honorary Assimilates'),
        ('11', 'Assimilated others'),
        ('99', 'Other regime'),
    ], string='Contracting Regime', required=True, default="02")
    payroll_type = fields.Selection([
            ('O', 'Ordinary Payroll'),
            ('E', 'Extraordinary Payroll')], 
            string='Payroll Type', 
            required=False,
            readonly=False,)
    payroll_month = fields.Selection([
            ('1', 'January'),
            ('2', 'February'),
            ('3', 'March'),
            ('4', 'April'),
            ('5', 'May'),
            ('6', 'June'),
            ('7', 'July'),
            ('8', 'August'),
            ('9', 'September'),
            ('10', 'October'),
            ('11', 'November'),
            ('12', 'December')], 
            string='Payroll month',
            required=True,
            readonly=True,
            states={'draft': [('readonly', False)]})
    payroll_of_month = fields.Selection([
            ('1', '1'),
            ('2', '2'),
            ('3', '3'),
            ('4', '4'),
            ('5', '5'),
            ('6', '6')], 
            string='Payroll of the month', 
            required=True, 
            default="1",
            readonly=True,
            states={'draft': [('readonly', False)]})
    payroll_period = fields.Selection([
            ('01', 'Daily'),
            ('02', 'Weekly'),
            ('10', 'Decennial'),
            ('04', 'Biweekly'),
            ('05', 'Monthly'),
            ('99', 'Otra Peridiocidad'),], 
            string='Payroll period', 
            default="04",
            required=True,
            readonly=True,
            states={'draft': [('readonly', False)]})
    table_id = fields.Many2one('table.settings', string="Table Settings")
    subtotal_amount_untaxed = fields.Float(string='Base imponible', readonly=True)
    amount_tax = fields.Float(string='Impuestos', readonly=True)
    payslip_count = fields.Integer(compute='_compute_payslip_count', string="Payslip Computation Details")
    payroll_tax_run_count = fields.Integer(compute='_compute_payroll_tax_run_count', string="Payslip Computation Details")
    state = fields.Selection([
        ('draft', 'Draft'),
        ('close', 'Close'),
        ('cancel', 'Cancelled'),
    ], string='Status', index=True, readonly=True, copy=False, default='draft')
    acumulated_amount_tax = fields.Float(string='Impuestos acumulados del mes')
    acumulated_subtotal_amount = fields.Float(string='Base imponible acumulada del mes')
    bonus_date = fields.Boolean('Bonus date', default=False)
    pay_bonus = fields.Boolean('Pay bonus?')
    pay_type = fields.Selection([('0','Efectivo'),('1','Especie')], string='Tipo de pago', default='0')
    tax_detail_lines = fields.One2many(inverse_name='payslip_run_id', comodel_name='hr.payroll.tax.details', string='Detalles de impuestos')
    employer_register_id = fields.Many2one('res.employer.register', "Employer Register", required=False, readonly=False)
    apply_honorarium = fields.Boolean('Cargar Honorarios?', default=False, readonly=True,
        states={'draft': [('readonly', False)]})
    iva_tax = fields.Float(required=True, digits=(16, 4), string='IVA', default=16.0000,
        readonly=True, states={'draft': [('readonly', False)]})
    iva_amount = fields.Float(required=True, digits=(16, 4), string='Importe del IVA',
        readonly=True, states={'draft': [('readonly', False)]})
    amount_honorarium = fields.Float(required=True, digits=(16, 4), string='Importe honorarios',
        readonly=True, states={'draft': [('readonly', False)]})
    apply_honorarium_on = fields.Selection([
        ('net', 'Total Neto'),
        ('gross', 'Total Bruto'),
    ], string='Aplicar honorarios sobre', index=True, copy=False,
        readonly=True, states={'draft': [('readonly', False)]})
    year = fields.Integer(string='Año', compute='_ge_year_period', store=True)
    generated = fields.Boolean('Generated', default=False)
    group_id = fields.Many2one('hr.group', string="Grupo/Empresa",readonly=True, states={'draft': [('readonly', False)]})
    payment_date = fields.Date(string='Fecha de pago', required=True,
        readonly=True, states={'draft': [('readonly', False)]})

    def get_data_bank(self):
        """ Function doc """
        line_data = []
        domain = [('slip_id.payslip_run_id','=', self.id), ('total','!=',0)]
        lines_ids = self.env['hr.payslip.line'].search(domain).filtered(lambda r: r.code == 'T001')
        for line in lines_ids:
            line_data.append({
                'employee_number': line.slip_id.employee_id.enrollment,
                'last_name': line.slip_id.employee_id.last_name,
                'mothers_last_name': line.slip_id.employee_id.mothers_last_name,
                'name': line.slip_id.employee_id.name,
                'bank_account': line.slip_id.employee_id.get_bank().bank_account if line.slip_id.employee_id.get_bank() else '',
                'total': line.total,
                'id_concept': '01 PAGO DE NOMINA',
            })
        return sorted(line_data, key=lambda k: k['employee_number'])

    def get_xls_bank(self):
        """ Function doc """
        output = io.BytesIO()
        file_xls = modules.get_module_resource('payroll_mexico', 'static/src/layout', 'Layout Dispersion Santander.xlsm')
        wb = load_workbook(filename = file_xls, read_only=False, keep_vba=True)
        ws1 = wb.active
        row = 7
        row += 1
        start_row = row
        payment_date = self.payment_date.strftime('%d/%m/%Y')
        ws1.cell(column=8, row=4, value=payment_date)
        for i, h in enumerate(self.get_data_bank()):
            j = i
            j += 1
            i += row
            # ~ ws1.cell(column=3, row=i, value=h.get('employee_number'))
            ws1.cell(column=3, row=i, value=j)
            ws1.cell(column=4, row=i, value=h.get('last_name'))
            ws1.cell(column=5, row=i, value=h.get('mothers_last_name'))
            ws1.cell(column=6, row=i, value=h.get('name'))
            ws1.cell(column=7, row=i, value=h.get('bank_account'))
            ws1.cell(column=8, row=i, value=h.get('total'))
            ws1.cell(column=9, row=i, value=h.get('id_concept'))
        row = i
        wb.save(output)
        f_name = 'Layout Dispersion Santander'
        xlsx_data = output.getvalue()
        export_id = self.env['hr.payslip.run.export.excel'].create({ 'excel_file': base64.encodestring(xlsx_data),'file_name': f_name + '.xlsm'})
        return {
            'view_mode': 'form',
            'res_id': export_id.id,
            'res_model': 'hr.payslip.run.export.excel',
            'view_type': 'form',
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    def get_data_bank_banorte(self):
        """ Function doc """
        line_data = []
        domain = [('slip_id.payslip_run_id','=', self.id), ('total','!=',0)]
        lines_ids = self.env['hr.payslip.line'].search(domain).filtered(lambda r: r.code == 'T001')
        for line in lines_ids:
            bank_account = line.slip_id.employee_id.get_bank().bank_account if line.slip_id.employee_id.get_bank() else ''
            line_data.append({
                'employee_number': line.slip_id.employee_id.enrollment.split('-')[1],
                'name': line.slip_id.employee_id.complete_name,
                'total': str(line.total).replace('.',''),
                'account_code': '072',
                'account_type': '001',
                'bank_account': bank_account.zfill(18),
            })
        return sorted(line_data, key=lambda k: k['employee_number'])

    def get_xls_bank_banorte(self):
        """ Function doc """
        output = io.BytesIO()
        create_date = self.create_date.strftime('%Y-%m-%d 23:59:59')
        sequence = self.search([('create_date','>=',self.create_date.date()),('create_date','<=',create_date)])
        file_xls = modules.get_module_resource('payroll_mexico', 'static/src/layout', 'Layout Dispersion Banorte.xlsm')
        wb = load_workbook(filename = file_xls, read_only=False, keep_vba=True)
        ws1 = wb.active
        row = 2
        row += 1
        payment_date = self.payment_date.strftime('%d/%-m/%Y')
        ws1.cell(column=8, row=2, value=payment_date)
        ws1.cell(column=9, row=2, value=len(sequence))
        total = 0
        for i, h in enumerate(self.get_data_bank_banorte()):
            total += float(h.get('total'))
            i += row
            ws1.cell(column=1, row=i, value=int(h.get('employee_number')))
            ws1.cell(column=2, row=i, value=h.get('name'))
            ws1.cell(column=3, row=i, value=float(h.get('total')))
            ws1.cell(column=4, row=i, value=int(h.get('account_code')))
            ws1.cell(column=5, row=i, value=int(h.get('account_type')))
            ws1.cell(column=6, row=i, value=h.get('bank_account'))
            
        row = i
        wb.save(output)
        f_name = 'Layout Dispersion Banorte'
        xlsx_data = output.getvalue()
        export_id = self.env['hr.payslip.run.export.excel'].create({ 'excel_file': base64.encodestring(xlsx_data),'file_name': f_name + '.xlsm'})
        return {
            'view_mode': 'form',
            'res_id': export_id.id,
            'res_model': 'hr.payslip.run.export.excel',
            'view_type': 'form',
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    @api.one
    @api.depends('date_start')
    def _ge_year_period(self):
        '''
        Este metodo obtiene el valor para el campo año basado en las fecha date_from de la nomina
        '''
        self.year = self.date_start.year

    def print_payslip_run_details(self):
        '''
        Este metodo imprime el reporte de payslip run
        '''
        return self.env.ref('payroll_mexico.report_payslip_run_template').report_action(self, {})

    @api.onchange('payroll_period','payroll_type')
    def payroll_period_extraordinary(self):
        if self.payroll_type == 'E':
            self.payroll_period = '99'
        if self.payroll_type == 'O':
            if self.payroll_period == '99':
                self.payroll_period = ''
            
    @api.onchange('apply_honorarium')
    def onchange_apply_honorarium(self):
        if not self.apply_honorarium:
            self.apply_honorarium_on = False
            self.amount_honorarium = False

    @api.onchange('apply_honorarium_on')
    def onchange_apply_honorarium_on(self):
        if self.apply_honorarium_on:
            self.set_tax_iva_honorarium()

    @api.onchange('iva_tax')
    def onchange_iva_tax(self):
        if self.iva_tax > 0:
            self.set_tax_iva_honorarium()
        else:
            self.iva_amount = 0

    def set_tax_iva_honorarium(self):
        domain = [('slip_id.payslip_run_id','=', self.id)]
        base_salary = 0
        neto = 0
        imss_rcv_infonavit = 0
        isr = 0
        for payroll in self:
            for slip in payroll.slip_ids:
                base_salary += sum(slip.line_ids.filtered(lambda l: l.code == 'P195').mapped('total'))
                neto += sum(slip.line_ids.filtered(lambda l: l.code == 'T001').mapped('total'))
                imss_rcv_infonavit += sum(slip.line_ids.filtered(lambda l: l.code in ('C001', 'D002', 'UI126', 'UI127', 'UI128')).mapped('total'))
                isr += sum(slip.line_ids.filtered(lambda l: l.code == 'D001').mapped('total'))
        isn = self.amount_tax
        if self.apply_honorarium:
            if self.group_id.percent_honorarium > 0:
                if self.apply_honorarium_on == 'net':
                    self.amount_honorarium = (self.group_id.percent_honorarium / 100) * neto
                if self.apply_honorarium_on == 'gross':
                    self.amount_honorarium = ((self.group_id.percent_honorarium / 100) * base_salary)
            else:
                raise ValidationError(_("Para Cargar 'Honorarios' a la nómina establesca el valor del porcentaje al registro del 'Grupo/Empresa'."))
        subtotal = neto + imss_rcv_infonavit + isr + isn + self.amount_honorarium
        iva = ((self.iva_tax / 100) * subtotal)
        if self.iva_tax > 0:
            self.iva_amount = iva

    def not_total(self):
        raise ValidationError(_('Nose encontraron valores para totalizar en la categoría NETO.'))

    
    @api.multi
    def generate_cfdi(self):
        for slip in self.slip_ids:
            slip.action_cfdi_nomina_generate()
        return
            
        
            
            
    
    @api.multi
    def print_payroll_summary_report(self):
        payroll_dic = {}
        employees = []
        total = 0
        domain = [('slip_id.payslip_run_id','=', self.id)]
        base_salary = sum(self.env['hr.payslip.line'].search(domain + [('code','=', 'P195')]).mapped('total'))
        neto = sum(self.env['hr.payslip.line'].search(domain + [('code','=', 'T001')]).mapped('total'))
        imss_rcv_infonavit = sum(self.env['hr.payslip.line'].search(domain + [('code','in', ('C001', 'D002', 'UI126', 'UI127', 'UI128'))]).mapped('total'))
        isr = sum(self.env['hr.payslip.line'].search(domain + [('code','=', 'D001')]).mapped('total'))
        isn = self.amount_tax
        honorarium = self.amount_honorarium
        subtotal = base_salary + imss_rcv_infonavit + isr + isn + honorarium
        iva = ((self.iva_tax / 100) * subtotal)
        total = subtotal + iva
        payroll_dic['regimen'] = self.contracting_regime
        payroll_dic['payroll_month'] = dict(self._fields['payroll_month']._description_selection(self.env)).get(self.payroll_month)
        payroll_dic['base_salary'] = base_salary
        payroll_dic['neto'] = neto
        payroll_dic['isn'] = isn
        payroll_dic['imss_rcv_infonavit'] = imss_rcv_infonavit
        payroll_dic['isr'] = isr
        
        payroll_dic['honorarium'] = honorarium
        payroll_dic['subtotal'] = subtotal
        payroll_dic['iva'] = iva
        payroll_dic['total'] = total
        data={
            'payroll_data':payroll_dic
            }
        return self.env.ref('payroll_mexico.action_payroll_summary_report').report_action(self,data)       

    @api.multi
    def print_payroll_deposit_report(self):
        payrolls = self.filtered(lambda s: s.state in ['close'])
        payroll_dic = {}
        employees = []
        total = 0
        for payroll in payrolls:
            payroll_dic['payroll_of_month'] = payroll.payroll_of_month
            payroll_dic['date_large'] = '%s a %s' %(payroll.date_start.strftime("%d/%b/%Y").title(), payroll.date_end.strftime("%d/%b/%Y").title())
            company = payroll.mapped('slip_ids').mapped('company_id')
            payroll_dic['rfc'] = company.rfc
            payroll_dic['employer_registry'] = company.employer_register_ids.filtered(lambda r: r.state == 'valid').mapped('employer_registry')[0] or ''
            for slip in payroll.slip_ids:
                total += sum(slip.line_ids.filtered(lambda r: r.category_id.code == 'NET').mapped('total'))
                employees.append({
                    'enrollment': slip.employee_id.enrollment,
                    'name': slip.employee_id.name_get()[0][1],
                    'bank_key': slip.employee_id.get_bank().bank_id.code if slip.employee_id.get_bank() else '',
                    'bank': slip.employee_id.get_bank().bank_id.name if slip.employee_id.get_bank() else '',
                    'account': slip.employee_id.get_bank().bank_account if slip.employee_id.get_bank() else '',
                    'total': slip.line_ids.filtered(lambda r: r.category_id.code == 'NET').mapped('total')[0] or self.not_total(),
                })
            payroll_dic['employees'] = employees
            payroll_dic['total_records'] = len(payroll.slip_ids)
        payroll_dic['total'] = total
        data={
            'payroll_data':payroll_dic
            }
        return self.env.ref('payroll_mexico.payroll_deposit_report_template').report_action(self,data)       

    @api.multi
    def print_fault_report(self):
        payroll_dic = {}
        payrolls = self.filtered(lambda s: s.state not in ['cancel'])
        leave_type = self.env['hr.leave.type'].search([('code','!=',False)])
        for payroll in payrolls:
            company = payroll.mapped('slip_ids').mapped('company_id')
            payroll_dic['rfc'] = company.rfc
            payroll_dic['date_start'] = '%s/%s/%s' %(payroll.date_start.strftime("%d"), payroll.date_start.strftime("%b").title(), payroll.date_start.strftime("%Y"))
            payroll_dic['date_end'] = '%s/%s/%s' %(payroll.date_end.strftime("%d"), payroll.date_end.strftime("%b").title(), payroll.date_end.strftime("%Y"))
            employee_ids = payroll.slip_ids.mapped('employee_id')
            fault_data = []
            for employee in employee_ids:
                
                for slip in payroll.slip_ids:
                    if employee.id == slip.employee_id.id:
                        total = 0
                        absenteeism = 0
                        inhability = 0
                        for leave in leave_type:
                            for wl in slip.worked_days_line_ids:
                                if leave.code == wl.code:
                                    # ~ total += wl.number_of_days
                                    if leave.time_type == 'inability':
                                        inhability += wl.number_of_days
                                    if leave.time_type == 'leave':
                                        absenteeism += wl.number_of_days
                        total += inhability + absenteeism
                        if total > 0:
                            fault_data.append({
                                'enrollment': employee.enrollment,
                                'name': employee.name_get()[0][1],
                                'fulltime': '---',
                                'total': round(total, 2),
                                'pay_company': '---',
                                '7mo': '---',
                                'inhability': round(inhability, 2),
                                'absenteeism': round(absenteeism, 2),
                            })
                payroll_dic['employee_data'] = fault_data
        
        data={
            'payroll_data': payroll_dic
            }
        return self.env.ref('payroll_mexico.action_fault_report').report_action(self,data)

    def _compute_acumulated_tax_amount(self):
        '''Este metodo calcula el impuesto acumulado para las nominas del mes'''
        current_year = fields.Date.context_today(self).year
        domain = [('group_id','=', self.group_id.id)]
        payslips_current_month = self.search([('payroll_month','=',self.payroll_month)] + domain).filtered(lambda sheet: sheet.date_start.year == current_year)
        total_tax_acumulated =  sum(payslips_current_month.mapped('amount_tax'))
        acumulated_subtotal_amount =  sum(payslips_current_month.mapped('subtotal_amount_untaxed'))
        payslips_current_month.write({'acumulated_amount_tax':total_tax_acumulated,
                                      'acumulated_subtotal_amount':acumulated_subtotal_amount})

    @api.multi
    def _compute_payslip_count(self):
        slip_ids = self.mapped('slip_ids')
        list_tax = []
        for payslip in slip_ids:
            line_ids = payslip.line_ids
            list_tax += line_ids.ids
        self.payslip_count = len(list_tax)

    @api.multi
    def _compute_payroll_tax_run_count(self):
        list_tax = []
        slip_ids = self.mapped('slip_ids')
        for payslip in slip_ids:
            line_ids = payslip.line_ids.filtered(lambda o: o.salary_rule_id.payroll_tax == True)
            list_tax += line_ids.ids
        self.payroll_tax_run_count = len(list_tax)
    
    @api.multi
    def action_view_payslip(self):
        return {
            'name': _('Detalles de Nómina'),
            # ~ 'domain': domain,
            'res_model': 'hr.payslip.line',
            'type': 'ir.actions.act_window',
            'view_id': False,
            'view_mode': 'tree,form',
            'view_type': 'form',
            'help': _('''<p class="oe_view_nocontent_create">
                           Click to Create for New Documents
                        </p>'''),
            'limit': 80,
        }
        
        return action
        
    @api.multi
    def action_view_payroll_tax_run(self):
        list_tax = []
        slip_ids = self.mapped('slip_ids')
        for payslip in slip_ids:
            line_ids = payslip.line_ids.filtered(lambda o: o.salary_rule_id.payroll_tax == True)
            list_tax += line_ids.ids
        domain = [('id', 'in', list_tax)]
        return {
            'name': _('Base Imp. ISN'),
            'domain': domain,
            'res_model': 'hr.payslip.line',
            'type': 'ir.actions.act_window',
            'view_id': False,
            'view_mode': 'tree,form',
            'view_type': 'form',
            'help': _('''<p class="oe_view_nocontent_create">
                           Click to Create for New Documents
                        </p>'''),
            'limit': 80,
        }
        return action
    
    @api.onchange('date_start', 'date_end','payroll_type')
    def onchange_date_start_date_end(self):
        if (not self.date_start) or (not self.date_end):
            return
        date_from = self.date_start
        date_to = self.date_end
        table_id = self.env['table.settings'].search([('year','=',int(date_from.year))],limit=1).id
        if not table_id:
            title = _("Aviso!")
            message = 'Debe configurar una tabla de configuracion para el año del periodo de la nómina.'
            warning = {
                'title': title,
                'message': message
            }
            self.update({'date_end': False})
            return {'warning': warning}
        self.table_id = self.env['table.settings'].search([('year','=',int(date_from.year))],limit=1).id
        self.payroll_month = str(date_from.month)
        date1 =datetime.strptime(str(str(date_from.year)+'-12-01'), DEFAULT_SERVER_DATE_FORMAT).date()
        date2 =datetime.strptime(str(str(date_from.year)+'-12-15'), DEFAULT_SERVER_DATE_FORMAT).date()
        
        if date_from >= date1 and date_to <= date2 and self.payroll_type == 'O':
            self.bonus_date = True
        else:
            self.bonus_date = False
        return
        
    @api.onchange('estructure_id')
    def onchange_estructure_id(self):
        if not self.estructure_id:
            return
        self.payroll_type = self.estructure_id.payroll_type
        return 

    @api.multi
    def compute_amount_untaxed(self):
        '''
        Este metodo calcula el monto de base imponible para la nomina a este monto se le calculara el impuesto
        '''
        self.slip_ids.compute_amount_untaxed()
        self.tax_detail_lines.unlink()
        dict_details={}
        list_details=[]
        for item in self.slip_ids.mapped('employee_id.work_center_id.id'):
            dict_details[item] = {
            'amount_untaxed': 0.0,
            'amount_tax': 0.0,
            }
        for slip in self.slip_ids:
            dict_details[slip.employee_id.work_center_id.id]['amount_untaxed'] += slip.subtotal_amount_untaxed
            dict_details[slip.employee_id.work_center_id.id]['amount_tax'] += slip.amount_tax
        for key in dict_details.keys():
            vals = {
                'work_center_id': key,
                'amount_untaxed': dict_details[key]['amount_untaxed'],
                'amount_tax':  dict_details[key]['amount_tax'],
                'payslip_run_id': self.id,
            }
            list_details.append(vals)
        self.tax_detail_lines = list_details
        self.subtotal_amount_untaxed = sum(self.tax_detail_lines.mapped('amount_untaxed'))
        self.amount_tax = sum(self.tax_detail_lines.mapped('amount_tax'))
        self._compute_acumulated_tax_amount()

    @api.multi
    def close_payslip_run(self):
        '''
        En este metodo se correran los calculos de base imponible e impuestos
        '''
        self.recalculate_payroll()
        self.slip_ids.compute_amount_untaxed()
        self.compute_amount_untaxed()
        for payslip in self.slip_ids:
            if not payslip.payment_date:
                raise ValidationError(_('Para poder cerrar la nómina debe agregar la Fecha de pago.'))
            payslip.state = 'done'
            amount = 0
            for line in payslip.line_ids:
                if line.salary_rule_id.type == 'deductions' and line.salary_rule_id.type_deduction == '011':
                    amount += line.total
            move_id = self.env['hr.credit.employee.account'].create_move(description=payslip.number,debit=amount,employee=payslip.employee_id)
            payslip.move_infonacot_id = move_id.id
            payslip.input_ids.write({'state':'paid'})
        return super(HrPayslipRun, self).close_payslip_run()
    
    @api.multi
    def draft_payslip_run(self):
        for payslip in self.slip_ids:
            payslip.state = 'draft'
            payslip.move_infonacot_id.unlink()
        return self.write({'state': 'draft'})
        
    @api.multi
    def cancel_payslip_run(self):
        for payslip in self.slip_ids:
            payslip.state = 'cancel'
            payslip.move_infonacot_id.unlink()
            payslip.input_ids.write({'payslip':False,'state':'approve'})
            payslip.input_ids = False
        return self.write({'state': 'cancel'})
    
    def recalculate_payroll(self):
        for payslip in self.slip_ids:
            # ~ vals={
                # ~ 'date_from':,
                # ~ 'date_to':,
                # ~ 'table_id':,
                # ~ '':,
                # ~ '':,
                # ~ '':,
            # ~ }
            worked_days_line_ids = payslip.get_worked_day_lines(payslip.contract_id, payslip.date_from, payslip.date_to)
            worked_days_lines = payslip.worked_days_line_ids.browse([])
            payslip.worked_days_line_ids = []
            for r in worked_days_line_ids:
                worked_days_lines += worked_days_lines.new(r)
            payslip.worked_days_line_ids = worked_days_lines
            payslip.compute_sheet()
        return 

class TaxDetails(models.Model):
    _name='hr.payroll.tax.details'
    _description='Detalles de impuestos para los procesamientos de nomina'

    #Columns
    work_center_id = fields.Many2one(comodel_name='hr.work.center',string='Centro de trabajo',required=False)
    amount_untaxed = fields.Float(string='Base Imponible',required=False)
    amount_tax = fields.Float(string='Impuestos',required=False)
    payslip_run_id = fields.Many2one(comodel_name='hr.payslip.run', string='Procesamiento de nomina')
